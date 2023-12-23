import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image

img_path = "example/baby Julia.jpg"
img_name, _ = os.path.splitext(img_path)
print(img_name)

num_rows, num_columns = 64, 64

def rgb_to_color(rgb_tuple):
    return f"FF{rgb_tuple[0]:02X}{rgb_tuple[1]:02X}{rgb_tuple[2]:02X}"

def remove_alpha(rgb_tuple):
    return rgb_tuple[:3] 

def apply_colors_from_rgb(rgb_list, sheet):
        # Set the column width and row height to make square cells
    for row_index in range(1, num_rows + 1):
        sheet.row_dimensions[row_index].height = 75/4

    for column_index in range(1, num_columns + 1):
        col_letter = get_column_letter(column_index)
        sheet.column_dimensions[col_letter].width = 12.43/4

    for row_index, row in enumerate(rgb_list, start=1):
        for col_index, rgb_color in enumerate(row, start=1):
            # Convert RGB to Color object
            rgb_color = remove_alpha(rgb_color)
            color = PatternFill(start_color=rgb_to_color(rgb_color), end_color=rgb_to_color(rgb_color), fill_type="solid")

            # Get the column letter
            col_letter = get_column_letter(col_index)

            # Apply the color to the cell
            sheet[f"{col_letter}{row_index}"].fill = color
                        # Debugging: Print RGB values for 'AD1'
            if col_index == 30 and row_index == 1:
                print(f"RGB for 'AD1': {rgb_color}")

img = Image.open(img_path)

imgSmall = img.resize((num_rows, num_columns))

# Scale back up using NEAREST to original size
result = imgSmall.resize(img.size,Image.NEAREST)

# Save
result.save(img_name+" pixelated.png")

pix_val = list(imgSmall.getdata())

rgb_list_2d = [pix_val[i:i+num_rows] for i in range(0, len(pix_val), num_rows)]

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Apply colors from the RGB list to the Excel sheet
apply_colors_from_rgb(rgb_list_2d, sheet)

# Save the workbook
workbook.save(img_name+'_excel.xlsx')
print('done')